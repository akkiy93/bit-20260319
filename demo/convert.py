"""
Markdown テスト仕様書 → Excel 変換スクリプト
使用方法: uv run python3 convert.py
"""

import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# ── 色設定 ──────────────────────────────────────────
COLOR_TITLE_BG   = "1F4E79"
COLOR_TITLE_FG   = "FFFFFF"
COLOR_H2_BG      = "2E75B6"
COLOR_H2_FG      = "FFFFFF"
COLOR_H3_BG      = "BDD7EE"
COLOR_H3_FG      = "1F4E79"
COLOR_HEADER_BG  = "4472C4"
COLOR_HEADER_FG  = "FFFFFF"
COLOR_ROW_ODD    = "FFFFFF"
COLOR_ROW_EVEN   = "DEEAF1"
COLOR_BORDER     = "B8CCE4"
COLOR_DONE_BG    = "D9D9D9"  # 完了行グレー
COLOR_DONE_FG    = "808080"  # 完了行テキスト

# ── チーム設定（ここを実際のメンバー名に変更） ──────────
DEFAULT_MEMBERS = [
    "担当者A",
    "担当者B",
    "担当者C",
    "担当者D",
    "担当者E",
]

# ── ステータス選択肢 ─────────────────────────────────
STATUS_OPTIONS = ["未着手", "進行中", "完了", "対象外"]

# ── スタイル生成 ─────────────────────────────────────
def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def apply_title(cell, text):
    cell.value = text
    cell.fill = make_fill(COLOR_TITLE_BG)
    cell.font = Font(bold=True, size=14, color=COLOR_TITLE_FG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_h2(cell, text):
    cell.value = text
    cell.fill = make_fill(COLOR_H2_BG)
    cell.font = Font(bold=True, size=12, color=COLOR_H2_FG)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def apply_h3(cell, text):
    cell.value = text
    cell.fill = make_fill(COLOR_H3_BG)
    cell.font = Font(bold=True, size=11, color=COLOR_H3_FG)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def apply_table_header(cell, text):
    cell.value = text
    cell.fill = make_fill(COLOR_HEADER_BG)
    cell.font = Font(bold=True, color=COLOR_HEADER_FG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()

def apply_table_cell(cell, text, row_index):
    cell.value = text
    color = COLOR_ROW_EVEN if row_index % 2 == 0 else COLOR_ROW_ODD
    cell.fill = make_fill(color)
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = make_border()

# ── 設定シートの作成 ─────────────────────────────────
def create_settings_sheet(wb, members):
    ws = wb.create_sheet("設定")

    # ヘッダー
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30

    header_cell = ws.cell(row=1, column=1, value="担当者リスト")
    header_cell.fill = make_fill(COLOR_HEADER_BG)
    header_cell.font = Font(bold=True, color=COLOR_HEADER_FG)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    note_cell = ws.cell(row=1, column=2, value="← 実際のメンバー名に変更してください")
    note_cell.font = Font(italic=True, color="808080")

    # メンバー名を入力
    for i, name in enumerate(members, start=2):
        cell = ws.cell(row=i, column=1, value=name)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = make_border()

    ws.row_dimensions[1].height = 22

    return len(members)  # メンバー数を返す

# ── Markdown パース ──────────────────────────────────
def parse_table(lines):
    rows = []
    for line in lines:
        line = line.strip()
        if not line.startswith("|"):
            break
        if re.match(r"^\|[-| :]+\|$", line):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        rows.append(cells)
    return rows

def parse_markdown(md_text):
    blocks = []
    lines = md_text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        if stripped.startswith("# ") and not stripped.startswith("## "):
            blocks.append({"type": "title", "content": stripped[2:].strip()})
            i += 1
        elif stripped.startswith("## "):
            blocks.append({"type": "h2", "content": stripped[3:].strip()})
            i += 1
        elif stripped.startswith("### "):
            blocks.append({"type": "h3", "content": stripped[4:].strip()})
            i += 1
        elif stripped.startswith("|"):
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                table_lines.append(lines[i])
                i += 1
            rows = parse_table(table_lines)
            if rows:
                blocks.append({"type": "table", "content": rows})
        else:
            i += 1

    return blocks

# ── Excel 書き込み ────────────────────────────────────
def write_excel(blocks, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "テスト仕様書"

    # 設定シートを作成し、メンバー数を取得
    member_count = create_settings_sheet(wb, DEFAULT_MEMBERS)
    # 設定シートの担当者リスト範囲（例: 設定!$A$2:$A$6）
    member_range = f"'設定'!$A$2:$A${1 + member_count}"

    col_widths = {}
    current_row = 1
    # ドロップダウン・条件付き書式の適用対象を記録
    # (start_row, end_row, status_col_idx, assignee_col_idx, total_cols)
    table_data_ranges = []

    def merge_and_write(apply_func, text, height=24):
        nonlocal current_row
        ws.row_dimensions[current_row].height = height
        cell = ws.cell(row=current_row, column=1, value=text)
        apply_func(cell, text)
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=12
        )
        current_row += 1

    for block in blocks:
        btype = block["type"]

        if btype == "title":
            merge_and_write(apply_title, block["content"], height=32)
            current_row += 1

        elif btype == "h2":
            current_row += 1
            merge_and_write(apply_h2, block["content"], height=22)

        elif btype == "h3":
            merge_and_write(apply_h3, block["content"], height=20)

        elif btype == "table":
            rows = block["content"]
            if not rows:
                continue

            header = rows[0]
            data_rows = rows[1:]

            # ステータス・担当者の列インデックスを検出（1始まり）
            status_col_idx    = None
            assignee_col_idx  = None
            for idx, h in enumerate(header, start=1):
                if h == "ステータス":
                    status_col_idx = idx
                elif h == "担当者":
                    assignee_col_idx = idx

            # ヘッダー行
            ws.row_dimensions[current_row].height = 22
            for col_idx, cell_text in enumerate(header, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                apply_table_header(cell, cell_text)
                col_letter = get_column_letter(col_idx)
                col_widths[col_letter] = max(
                    col_widths.get(col_letter, 0), len(cell_text) + 4
                )
            current_row += 1

            # データ行
            data_start_row = current_row
            for row_i, row in enumerate(data_rows):
                ws.row_dimensions[current_row].height = 40
                for col_idx, cell_text in enumerate(row, start=1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    apply_table_cell(cell, cell_text, row_i)
                    col_letter = get_column_letter(col_idx)
                    col_widths[col_letter] = max(
                        col_widths.get(col_letter, 0),
                        min(len(cell_text), 50)
                    )
                current_row += 1
            data_end_row = current_row - 1

            # ドロップダウン適用対象として記録
            if data_rows and (status_col_idx or assignee_col_idx):
                table_data_ranges.append({
                    "start": data_start_row,
                    "end":   data_end_row,
                    "status_col":   status_col_idx,
                    "assignee_col": assignee_col_idx,
                    "total_cols":   len(header),
                })

    # ── 列幅の適用 ──────────────────────────────────────
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = max(width * 1.8, 10)

    # ── プルダウン・条件付き書式の適用 ──────────────────
    for tr in table_data_ranges:
        start = tr["start"]
        end   = tr["end"]
        total = tr["total_cols"]
        end_col_letter = get_column_letter(total)

        # ステータス プルダウン
        if tr["status_col"]:
            status_letter = get_column_letter(tr["status_col"])
            dv_status = DataValidation(
                type="list",
                formula1=f'"{",".join(STATUS_OPTIONS)}"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="入力エラー",
                error=f'選択肢から選んでください：{" / ".join(STATUS_OPTIONS)}',
            )
            dv_status.sqref = f"{status_letter}{start}:{status_letter}{end}"
            ws.add_data_validation(dv_status)

            # 条件付き書式：ステータスが「完了」の場合、行全体をグレーアウト
            done_fill = PatternFill(bgColor=COLOR_DONE_BG)
            done_font = Font(color=COLOR_DONE_FG)
            dxf_done = DifferentialStyle(fill=done_fill, font=done_font)
            rule_done = Rule(
                type="expression",
                dxf=dxf_done,
                formula=[f'${status_letter}{start}="完了"'],
            )
            ws.conditional_formatting.add(
                f"A{start}:{end_col_letter}{end}", rule_done
            )

            # 条件付き書式：ステータスが「対象外」の場合も薄くグレーアウト
            skip_fill = PatternFill(bgColor="EFEFEF")
            skip_font = Font(color="AAAAAA", italic=True)
            dxf_skip = DifferentialStyle(fill=skip_fill, font=skip_font)
            rule_skip = Rule(
                type="expression",
                dxf=dxf_skip,
                formula=[f'${status_letter}{start}="対象外"'],
            )
            ws.conditional_formatting.add(
                f"A{start}:{end_col_letter}{end}", rule_skip
            )

        # 担当者 プルダウン（設定シート参照）
        if tr["assignee_col"]:
            assignee_letter = get_column_letter(tr["assignee_col"])
            dv_assignee = DataValidation(
                type="list",
                formula1=member_range,
                allow_blank=True,
                showErrorMessage=False,  # 自由入力も許可
            )
            dv_assignee.sqref = f"{assignee_letter}{start}:{assignee_letter}{end}"
            ws.add_data_validation(dv_assignee)

    ws.freeze_panes = None
    wb.save(output_path)

    print(f"✅ 変換完了: {output_path}")
    print(f"   テスト仕様書シート: {current_row - 1} 行 × {len(col_widths)} 列")
    print(f"   設定シート: 担当者 {member_count} 名")
    print(f"   プルダウン適用テーブル: {len(table_data_ranges)} 個")


# ── メイン ───────────────────────────────────────────
if __name__ == "__main__":
    input_path  = Path(__file__).parent / "test_spec.md"
    output_path = Path(__file__).parent / "test_spec.xlsx"

    if not input_path.exists():
        print(f"❌ ファイルが見つかりません: {input_path}")
        exit(1)

    md_text = input_path.read_text(encoding="utf-8")
    blocks  = parse_markdown(md_text)
    write_excel(blocks, output_path)
